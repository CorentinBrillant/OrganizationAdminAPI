from copy import copy
from io import BytesIO
from pathlib import Path

from cryptography.fernet import Fernet, InvalidToken
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PillowImage
from PIL import UnidentifiedImageError

from ..models import Campaign, Member
from .helloasso_service import HelloAssoAPIError, HelloAssoConfigError, HelloAssoService

TEMPLATE_FILENAME = "Fichier Badges Adhérent 2025-2026 CKCP (1).xlsx"
PHOTO_WIDTH = 130
PHOTO_HEIGHT = 130


class BadgeOrderExportError(Exception):
    """Raised when the badge-order workbook cannot be generated."""


def _valid_image_content(content: bytes | None) -> bytes | None:
    if not content:
        return None
    try:
        with PillowImage.open(BytesIO(content)) as image:
            image.verify()
    except (OSError, UnidentifiedImageError):
        return None
    return content


def _ffck_photo_content(member: Member, *, current_photo_only: bool = False) -> bytes | None:
    rows = sorted(
        (row for row in member.ffck_export_rows.all() if row.photo),
        key=lambda row: (row.ffck_export.fetched_at, row.id),
        reverse=True,
    )
    for row in rows:
        if current_photo_only and member.photo != f"/api/ffck/rows/{row.id}/photo/download/":
            continue

        try:
            with row.photo.open("rb") as photo_file:
                content = photo_file.read()
        except FileNotFoundError:
            return None

        if not row.photo.name.endswith(".enc"):
            return _valid_image_content(content)

        key = str(getattr(settings, "MEMBER_CERTIFICAT_ENCRYPTION_KEY", "")).strip()
        if not key:
            return None
        try:
            return _valid_image_content(Fernet(key.encode("utf-8")).decrypt(content))
        except (InvalidToken, ValueError):
            continue
    return None


def _photo_content_for_member(
    member: Member, helloasso_service: HelloAssoService | None
) -> bytes | None:
    if not member.photo:
        return _ffck_photo_content(member)

    local_photo = _ffck_photo_content(member, current_photo_only=True)
    if local_photo is not None:
        return local_photo

    if helloasso_service is None:
        return _ffck_photo_content(member)

    try:
        document = helloasso_service.download_document(member.photo)
        helloasso_photo = _valid_image_content(document.content)
    except (HelloAssoConfigError, HelloAssoAPIError):
        helloasso_photo = None
    return helloasso_photo or _ffck_photo_content(member)


def _copy_template_row(sheet, source_row: int, target_row: int) -> None:
    for column in range(1, 12):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _add_photo(sheet, content: bytes | None, cell_reference: str) -> None:
    if not content:
        return
    try:
        image = Image(BytesIO(content))
    except Exception:
        return
    image.width = PHOTO_WIDTH
    image.height = PHOTO_HEIGHT
    sheet.add_image(image, cell_reference)


class BadgeOrderExportService:
    def __init__(self, campaign: Campaign):
        self.campaign = campaign
        self._helloasso_service: HelloAssoService | None = None

    def _get_helloasso_service(self) -> HelloAssoService:
        if self._helloasso_service is None:
            self._helloasso_service = HelloAssoService(
                client_id=getattr(settings, "HELLOASSO_CLIENT_ID", "").strip(),
                client_secret=getattr(settings, "HELLOASSO_CLIENT_SECRET", "").strip(),
            )
        return self._helloasso_service

    def export(self) -> bytes:
        template_path = Path(settings.BASE_DIR) / "assets" / TEMPLATE_FILENAME
        if not template_path.is_file():
            raise BadgeOrderExportError(f"Badge template file not found: {TEMPLATE_FILENAME}")

        try:
            workbook = load_workbook(template_path)
        except Exception as exc:
            raise BadgeOrderExportError("Unable to read badge template workbook.") from exc

        sheet = workbook.active
        max_template_row = sheet.max_row
        sheet._images = []
        for row in sheet.iter_rows(min_row=3, max_row=max_template_row, max_col=11):
            for cell in row:
                cell.value = None

        members = [
            member
            for member in Member.objects.filter(campaign=self.campaign).prefetch_related(
                "ffck_export_rows__ffck_export"
            )
            if not member.is_deleted
        ]
        members.sort(
            key=lambda member: (
                member.name.casefold(),
                member.first_name.casefold(),
                member.id,
            )
        )

        for index, member in enumerate(members):
            row_number = 3 + index // 2
            if row_number > max_template_row:
                _copy_template_row(sheet, 3, row_number)

            start_column = 1 if index % 2 == 0 else 7
            sheet.cell(row_number, start_column).value = index + 1
            sheet.cell(row_number, start_column + 1).value = member.name.upper()
            sheet.cell(row_number, start_column + 2).value = member.first_name
            sheet.cell(row_number, start_column + 3).value = None
            photo_cell = sheet.cell(row_number, start_column + 4)
            helloasso_service = None
            if str(member.photo or "").startswith(("https://", "http://")):
                helloasso_service = self._get_helloasso_service()
            _add_photo(
                sheet,
                _photo_content_for_member(member, helloasso_service),
                photo_cell.coordinate,
            )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
